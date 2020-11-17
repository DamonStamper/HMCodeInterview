input_filename = 'Sample A.xlsx'
output_filename = 'Sample A - Output.csv'

logging_level = 'DEBUG'
try:
    import logging
    import datetime

    import openpyxl
    import pandas
    import helper_save
    import helper_wrappers
except:
    raise Exception("Could not load required python libraries. Please run 'pip install -r requirements.txt' then try again.")

#Set logging options.
logging.basicConfig(format='%(asctime)s %(message)s')
logger = logging.getLogger(__name__)
logging_level = logging_level.upper()
loglevels = {
    'CRITICAL' : logging.CRITICAL,
    'ERROR' : logging.ERROR,
    'WARNING' : logging.WARNING,
    'INFO' : logging.INFO,
    'DEBUG' : logging.DEBUG
}
level = loglevels[logging_level]
logger.setLevel(level)

@helper_wrappers.callLogger
@helper_wrappers.timer
def main():
    data = getData(input_filename)
    data = cleanData(data)
    data = formatDataForSaving(data) # Doing this just before time since there is some data loss (rounding numbers) and may cause unexpected results otherwise.
    helper_save.saveData(data, output_filename)

@helper_wrappers.callLogger
@helper_wrappers.timer
def getData(input_filename): # non-OOP adapter pattern
    data = getDataFromExcel(input_filename)
    return data

@helper_wrappers.callLogger
@helper_wrappers.timer
def getDataFromExcel(input_filename):
    workbook = pandas.read_excel(input_filename, header=5, dtype=object) # This sets the column labels and removes the header(first 5 rows)
    workbook = addExtraColumnFromExcel(input_filename, workbook)
    logger.debug(f'Column labels:\n {workbook.columns.tolist()}')
    return workbook

@helper_wrappers.callLogger
@helper_wrappers.timer
def addExtraColumnFromExcel(input_filename, workbook):
    wb = openpyxl.load_workbook(input_filename)
    ws = wb.active
    column_addendum_header = ws['A1'].value
    logger.debug(f'Additional column header to add:\n{column_addendum_header}')
    column_addendum_values = (
        ws['A2'].value,
        ws['A3'].value,
        ws['A4'].value
        )
    column_addendum_value = "\n".join(column_addendum_values)
    logger.debug(f'Additional column value to add:\n{column_addendum_value}')
    workbook[column_addendum_header] = column_addendum_value
    return workbook

@helper_wrappers.callLogger
@helper_wrappers.timer
def cleanData(data):
    data = FillInMissingData(data)
    invalidGroupValues = ['Additional Notice                                                                                                         Test']
    data = data[~data['Group'].isin(invalidGroupValues)]
    asciiMask = list(map(isSeriesascii, data['Group']))
    data = data[asciiMask]
    return data

def isSeriesascii(s):
    return s.isascii()

@helper_wrappers.callLogger
@helper_wrappers.timer
def FillInMissingData(data):
    data = data.ffill() # Using ffill to propagate data from "top" rows to "below" rows because input data was designed for human readers in that they would assume missing data on "child" rows could be found on "parent" rows.
    return data

@helper_wrappers.callLogger
@helper_wrappers.timer
def formatDataForSaving(data):
    data = formatTotalRows(data)
    data = formatDateTime(data)
    return data

@helper_wrappers.callLogger
@helper_wrappers.timer
def formatDateTime(data):
    troublesomeTimeColumns = ('Finalized\nDate','Service\nDate From','Service\nDate To')
    for column in troublesomeTimeColumns:
        data[column] = list(map(dateFix, data[column]))

    troublesomeCurrencyColumns = {
        'Allowance': '${:,.2f}',
        'Paid\nAmount': '${:,.2f}'
        }
    for key, value in troublesomeCurrencyColumns.items():
        data[key] = data[key].apply(value.format)
        data[key] = list(map(currencyFixNegativeValues, data[key]))
    return data

# @helper_wrappers.callLogger
# @helper_wrappers.timer
def currencyFixNegativeValues(input):
    if '$-' in input:
        # logger.debug('Replacing $- with -$')
        input = input.replace('$-','-$')
        # logger.debug(f'Result: {input}')
    return input

# @helper_wrappers.callLogger
# @helper_wrappers.timer
def dateFix(input):
    iterant = pandas.to_datetime(input, errors='ignore')
    try:
        iterant = iterant.strftime('%m/%d/%Y')
    except Exception:
        pass
    return iterant

@helper_wrappers.callLogger
@helper_wrappers.timer
def formatTotalRows(dataframe):
    #due to the possibly dynamic nature of cell A1 ("TEST") changing we have to figure out what the column names are--and then create a dict that can be inserted as a blank series after total so that we match expected output formatting.
    columns = dataframe.columns.tolist()
    columnDict = {}
    for column in columns:
        columnDict[column] = ''

    indexOfTotalRows = dataframe.index[dataframe['Group'] == 'Total'].tolist()
    logger.debug(f'Indeces of total rows: {indexOfTotalRows}')
    dataframe = cleanTotalsColumns(dataframe, columnDict, indexOfTotalRows)
    dataframe = insertBlankSeriesAfterIndex(dataframe, columnDict, indexOfTotalRows)
    dataframe = dataframe[:-2] # remove last 2 rows from dataset (the 2nd total, and the extra row we added as part of insertBlankSeriesAfterIndex()
    return dataframe

@helper_wrappers.callLogger
@helper_wrappers.timer
def insertBlankSeriesAfterIndex(dataframe, columnDict, rowIndices):
    line = pandas.DataFrame(columnDict, index=rowIndices)
    dataframe = dataframe.append(line, ignore_index=False)
    dataframe = dataframe.sort_index().reset_index(drop=True)
    return dataframe

@helper_wrappers.callLogger
@helper_wrappers.timer
def cleanTotalsColumns(dataframe, columnDict, indexOfTotalRows):
    columnDataToKeep = ('Group', 'Allowance', 'Paid\nAmount')
    for column in columnDataToKeep:
        del columnDict[column]
    logger.debug(f'Remove data from these columns on Total lines:\n{columnDict}')

    for rowIndex in indexOfTotalRows:
        for columnLabel in columnDict:
            dataframe.at[rowIndex, columnLabel] = ''
    return dataframe

main()