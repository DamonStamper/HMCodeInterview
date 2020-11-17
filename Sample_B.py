input_filename = 'Sample B.xlsx'
output_filename = 'Sample B - Output.csv'

columnInfo = {
    'name': 'A4', 
    'IDNumber': 'A5', 
    'contractPeriod': 'I3',
    'contractBasis': 'I4',
    'Paid Claims Date': 'I5',
    'type': 'I6'
    }

logging_level = 'DEBUG'
try:
    import logging
    import platform

    import openpyxl
    import pandas
    import numpy

    import helper_save
    import helper_wrappers
except:
    raise Exception("Could not load required python libraries. Please run 'pip install -r requirements.txt' then try again.")

#Set logging options.
logging.basicConfig(format='%(asctime)s %(message)s')
logger = logging.getLogger(__name__)
logging_level = logging_level.upper()
loglevels = {
    'CRITICAL' : logging.CRITICAL,
    'ERROR' : logging.ERROR,
    'WARNING' : logging.WARNING,
    'INFO' : logging.INFO,
    'DEBUG' : logging.DEBUG
}
level = loglevels[logging_level]
logger.setLevel(level)

@helper_wrappers.callLogger
@helper_wrappers.timer
def main():
    data = getData(input_filename)
    data = sumData(data)
    data = formatDataForSaving(data) # Doing this just before time since there is some data loss (rounding numbers) and may cause unexpected results otherwise.
    helper_save.saveData(data, output_filename)

@helper_wrappers.callLogger
@helper_wrappers.timer
def sumData(dataframe):
    """Return a dataframe with an additional row containing the sum of rows contained in the list columnsToSum"""
    # TODO: Better way to do this https://stackoverflow.com/a/62734983/7902967
    columnsToSum = ['CHARGES', 'OOP', 'ACCESS_FEES', 'SPECL_DED', 'COPAY', 'NON_COVERED', 'BYD', 'MM_PAY', 'OOA_DRG', 'ITS_SURCHARGE','TOTAL']
    sum = dataframe.sum(axis = 0, skipna = True)
    dataframe = dataframe.append(sum[columnsToSum], ignore_index=True)
    logger.debug(sum[columnsToSum])
    dataframe = dataframe.round(2)
    dataframe['PRIM_PVDR_NO'].iloc[-1] = 'TOTAL' #Set last row's "PRIM_PVDR_NO" column to 'TOTAL'
    return dataframe

@helper_wrappers.callLogger
@helper_wrappers.timer
def getData(input_filename): # non-OOP adapter pattern
    """'interface' for getting cleaned/scrubbed data from a source"""
    data = getDataFromExcel(input_filename)
    return data

@helper_wrappers.callLogger
@helper_wrappers.timer
def getDataFromExcel(input_filename):
    """Implementation of getData for Excel (xls, xlsx) sources. Returns cleaned dataframe."""
    EnrollmentInformation = pandas.read_excel(input_filename, sheet_name = 'Enrollment Information', header=0, dtype=object)
    Claims = pandas.read_excel(input_filename, sheet_name = 'Claims 02-13-20', header=9, converters= {'DTE_DISP':pandas.to_datetime, 'DTE_SRVC_BEG':pandas.to_datetime, 'DTE_SRVC_END':pandas.to_datetime, 'PAT_ID':int, 'YTD Total Amount ':int, 'Reimbursement \nAmt. Requested':int})
    logger.info('Please note that expected output calls for "YTD Total Amount" and "Reimbursement Amt. Requested" fields to be integers (which means no decimal values). However this is currency which means that we are not dealing in whole units. I am going to make a judgement call and allow these fields to be floats/have decimal values.')
    additonalColumnsDataframe = claimsExtraInfo()

    #Store column labels for later reference during formatDataForSaving()
    global EnrollmentInformationColumns
    EnrollmentInformationColumns = EnrollmentInformation.columns.tolist()
    global ClaimsColumns
    ClaimsColumns = Claims.columns.tolist()
    global additonalColumnsDataframeColumns
    additonalColumnsDataframeColumns = additonalColumnsDataframe.columns.tolist()

    dataframe = mergeDataframes(EnrollmentInformation, additonalColumnsDataframe)
    dataframe = mergeDataframes(dataframe, Claims)
    dataframe = fillDataframeDesiredData(dataframe)
    dataframe = removePaddedZeros(dataframe, ["CODE 1"])
    dataframe = fillDataframeFromTo(EnrollmentInformation, dataframe)
    dataframe = removeTotals(dataframe)
    return dataframe

@helper_wrappers.callLogger
@helper_wrappers.timer
def removeTotals(dataframe):
    """Data cleanup by removing rows containing 'TOTAL'."""
    # TODO: Pull this out into an interface/implementation.
    dataframe = dataframe.drop(dataframe[dataframe['PRIM_PVDR_NO'] == 'TOTAL'].index)
    return dataframe

@helper_wrappers.callLogger
@helper_wrappers.timer
def removePaddedZeros(dataframe, columns):
    """Data cleanup by removing left padded zeros"""
    dataframe[columns] = dataframe[columns].apply(pandas.to_numeric)
    return dataframe

@helper_wrappers.callLogger
@helper_wrappers.timer
def fillDataframeDesiredData(dataframe):
    """Data cleanup by forward filling columns in the cols list. Returns dataframe.
    Forward filling means cell A1 will will A2 if A2 doesn't have data, if A2 has data then A2 will be used to fill A3 if it doesn't have data, etc    """
    cols = list(columnInfo.keys())
    logger.debug(f'cols to ffill():\n{cols}')
    dataframe.loc[:,cols] = dataframe.loc[:,cols].ffill()
    return dataframe

@helper_wrappers.callLogger
@helper_wrappers.timer
def fillDataframeFromTo(fromDataframe, toDataframe):
    """Append to columns/labels in 'toDataframe' with columns/labels from 'fromDataframe', resulting in toDataframe's columns to be the leftmost columns.
    Forward fills the resulting dataframe."""
    cols = fromDataframe.columns.tolist()
    toDataframe.loc[:,cols] = toDataframe.loc[:,cols].ffill()
    return toDataframe

# @helper_wrappers.callLogger
# @helper_wrappers.timer
def dateFix(input):
    """Format the input variable to a non 0-padded 'month/day/year hour:minute' format"""
    # TODO: Is running platform.system() performant given that this function is called repeatedly via map? system() means this is a method, but is it just an accessor method or something that does logic that doesn't need repeated ROWS*3 times?
    iterant = pandas.to_datetime(input, errors='ignore')
    try:
        if platform.system() != 'Windows':
            iterant = iterant.strftime('%-m/%-d/%Y %-H:%M')
        else:
            iterant = iterant.strftime('%#m/%#d/%Y %#H:%M')
    except Exception:
        pass
    return iterant

@helper_wrappers.callLogger
@helper_wrappers.timer
def setDateTimeColumns(data):
    """Sets the columns in datetimeColumns to a specific datetime format."""
    datetimeColumns = ('DTE_DISP','DTE_SRVC_BEG','DTE_SRVC_END')
    for column in datetimeColumns:
        logger.debug(f'Converting datetime in {column}')
        data[column] = list(map(dateFix, data[column]))
        # data[column]
    return data

@helper_wrappers.callLogger
@helper_wrappers.timer
def claimsExtraInfo():
    """Returns dataframe from 'claims *' worksheet. Due to date being included in the sheet name this works on a match basis and may cause issues if multiple sheets contain the word 'claims'"""
    # TODO: This feels hardcoded to excel spreadsheets. Abstract this out to something better?
    wb = openpyxl.load_workbook(input_filename)
    for s in range(len(wb.sheetnames)):
        if 'Claims' in  wb.sheetnames[s]:
            wb.active = s
            ws = wb.active
            logger.debug(ws.title)

    # Kindof a reverse interpolation? Replace the values in the dict with the appropriate, cleaned, data in the spreadsheet thus creating a key value pair representing column name and column value.
    for columnName, location in columnInfo.items():
        columnInfo[columnName] = (ws[location].value).split(':')[1].strip()

    dataframe = pandas.DataFrame([columnInfo], columns=columnInfo.keys())
    return dataframe

@helper_wrappers.callLogger
@helper_wrappers.timer
def mergeDataframes(dataframe1, dataframe2):
    """Combines two dataframes. Does not do any advanced processing (ffill, removing 0 pads, etc)"""
    mergedDataframe = dataframe2.join(dataframe1[dataframe1.columns])
    return mergedDataframe

@helper_wrappers.callLogger
@helper_wrappers.timer
def formatDataForSaving(dataframe):
    """Arrange columns in specific order"""
    cols = (EnrollmentInformationColumns + additonalColumnsDataframeColumns + ClaimsColumns)
    dataframe = dataframe[cols]
    dataframe = setDateTimeColumns(dataframe)
    return dataframe

@helper_wrappers.callLogger
@helper_wrappers.timer
def convertToIntWhereWhole(dataframe):
    def change_dtype(value):
        try:
            return int(value)
        except ValueError:
            try:
                return float(value)
            except ValueError:
                return value

    for column in dataframe.columns:
        dataframe.loc[:, column] = dataframe[column].apply(change_dtype)
    return dataframe

main()